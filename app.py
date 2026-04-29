from flask import Flask, render_template, request, redirect, session, jsonify, send_from_directory
import pandas as pd
import sqlite3
import requests
import io
import re
import gspread
from google.oauth2.service_account import Credentials
from datetime import datetime, timedelta
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload
from flask_mail import Mail, Message
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from apscheduler.schedulers.background import BackgroundScheduler
import atexit
from io import BytesIO
import os
import json
import threading
import psycopg2
import psycopg2.extras



app = Flask(__name__)
app.secret_key = "segredo123"

app.config['MAIL_SERVER']         = 'smtp.gmail.com'
app.config['MAIL_PORT']           = 465
app.config['MAIL_USE_TLS']        = True
app.config['MAIL_USE_SSL']        = False
app.config['MAIL_USERNAME']       = 'notificacoes.sistema@edusjc.sp.gov.br'
app.config['MAIL_PASSWORD']       = 'ubvw axxm azyt kgka'
app.config['MAIL_DEFAULT_SENDER'] = ('Sistema CITE', 'notificacoes.sistema@edusjc.sp.gov.br')

mail = Mail(app)

FUNCIONARIOS_URL = "https://docs.google.com/spreadsheets/d/1nYgCV6SgPf5PTIIJZLQcM9xfw5L08SNber7ZEDfUWOo/export?format=csv"

DATABASE_URL = os.environ.get("DATABASE_URL")

# =========================
# 🛠️ CONEXÃO COM BANCO
# =========================

def get_conn():
    """Retorna conexão com PostgreSQL (produção) ou SQLite (local)."""
    if DATABASE_URL:
        conn = psycopg2.connect(DATABASE_URL)
        return conn
    else:
        return sqlite3.connect("kanban.db")

def is_pg():
    return DATABASE_URL is not None

def ph():
    """Retorna o placeholder correto: %s para PG, ? para SQLite."""
    return "%s" if is_pg() else "?"

def fix_sql(sql):
    """Converte placeholders ? para %s se estiver usando PostgreSQL."""
    if is_pg():
        return sql.replace("?", "%s")
    return sql

def cursor(conn):
    """Retorna cursor adequado para o banco."""
    if is_pg():
        return conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    return conn.cursor()

def fetchall_as_dicts(c, keys):
    """Converte fetchall do SQLite para lista de dicts (PG já retorna dicts)."""
    if is_pg():
        rows = c.fetchall()
        return [dict(r) for r in rows]
    return [dict(zip(keys, r)) for r in c.fetchall()]

# =========================
# 🛠️ UTILITÁRIOS
# =========================

def extrair_url_editavel(url_csv):
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url_csv)
    if match:
        sheet_id = match.group(1)
        return f"https://docs.google.com/spreadsheets/d/{sheet_id}/edit"
    return None


def get_gspread_client():
    scope = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]
    cred_json = os.environ.get("GOOGLE_CREDENTIALS")
    cred_dict = json.loads(cred_json)

    creds = Credentials.from_service_account_info(cred_dict, scopes=scope)
    return gspread.authorize(creds)

def extrair_sheet_e_gid(url):
    sheet_match = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url)
    gid_match = re.search(r"gid=(\d+)", url)
    if not sheet_match:
        raise Exception("URL inválida — ID da planilha não encontrado.")
    sheet_id = sheet_match.group(1)
    gid = gid_match.group(1) if gid_match else "0"
    return sheet_id, gid


def abrir_aba_por_gid(sheet_id, gid):
    client = get_gspread_client()
    sh = client.open_by_key(sheet_id)
    for ws in sh.worksheets():
        if str(ws.id) == str(gid):
            return ws
    raise Exception(f"Aba com gid={gid} não encontrada.")


def get_url_sgrh_funcionario(nome_funcionario):
    try:
        df = pd.read_csv(FUNCIONARIOS_URL)
        df.columns = df.columns.str.strip()
        df = df.fillna("")
        df["NOME"] = df["NOME"].astype(str).str.strip()
        row = df[df["NOME"] == nome_funcionario]
        if row.empty:
            return None
        url = str(row.iloc[0].get("PLANILHA SGRH", "")).strip()
        return url if url else None
    except Exception as e:
        print("ERRO get_url_sgrh:", e)
        return None


LABELS_JUSTIFICATIVA = {
    "95":    "95 - SERVIÇO EXTERNO",
    "7":     "7 -ABONADA",
    "150":   "150 - TRE",
    "90":    "90 - JUSTIFICADO COM AUTORIZAÇÃO DA CHEFIA",
    "outro": "OUTRO",
    "troca": "TROCA DE HORÁRIO",
    "":      "",
}
LABELS_PARA_VALOR = {v.strip().upper(): k for k, v in LABELS_JUSTIFICATIVA.items() if v}

LINHA_Q1 = 10
LINHA_Q2 = 31
LINHAS_QUINZENA = 13


def ler_quinzena(ws, linha_inicio):
    fim = linha_inicio + LINHAS_QUINZENA - 1
    valores = ws.get(f"B{linha_inicio}:E{fim}")
    linhas = []
    for i in range(LINHAS_QUINZENA):
        row = valores[i] if i < len(valores) else []
        data    = row[0].strip() if len(row) > 0 else ""
        hora    = row[1].strip() if len(row) > 1 else ""
        just_lb = row[2].strip().upper() if len(row) > 2 else ""
        obs     = row[3].strip() if len(row) > 3 else ""
        linhas.append({"data": data, "horario": hora, "just": LABELS_PARA_VALOR.get(just_lb, ""), "obs": obs})
    return linhas


def salvar_quinzena(ws, linha_inicio, linhas):
    dados = []
    for l in linhas[:LINHAS_QUINZENA]:
        dados.append([
            l.get("data", ""),
            l.get("horario", ""),
            LABELS_JUSTIFICATIVA.get(l.get("just", ""), ""),
            l.get("obs", ""),
        ])
    while len(dados) < LINHAS_QUINZENA:
        dados.append(["", "", "", ""])
    fim = linha_inicio + LINHAS_QUINZENA - 1
    ws.update(f"B{linha_inicio}:E{fim}", dados, value_input_option="USER_ENTERED")

def criar_notificacao(conn, destinatario, tipo, mensagem, link=None):
    c = cursor(conn)
    c.execute(fix_sql("""
        INSERT INTO notificacoes (destinatario, tipo, mensagem, link)
        VALUES (?, ?, ?, ?)
    """), (destinatario, tipo, mensagem, link))


# =========================
# 🧠 DB
# =========================

def init_db():
    conn = get_conn()
    c = cursor(conn)

    serial = "SERIAL" if is_pg() else "INTEGER"
    autoincrement = "" if is_pg() else "AUTOINCREMENT"

    tables = [
        f"""CREATE TABLE IF NOT EXISTS tasks (
            id {serial} PRIMARY KEY {autoincrement},
            nome TEXT,
            titulo TEXT,
            descricao TEXT,
            status TEXT DEFAULT 'todo'
        )""",
        f"""CREATE TABLE IF NOT EXISTS comunicados (
            id {serial} PRIMARY KEY {autoincrement},
            titulo TEXT,
            texto TEXT,
            autor TEXT,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        f"""CREATE TABLE IF NOT EXISTS reacoes (
            id {serial} PRIMARY KEY {autoincrement},
            comunicado_id INTEGER,
            usuario TEXT,
            emoji TEXT,
            UNIQUE(comunicado_id, usuario, emoji)
        )""",
        f"""CREATE TABLE IF NOT EXISTS eventos (
            id {serial} PRIMARY KEY {autoincrement},
            titulo TEXT,
            data TEXT,
            tipo TEXT DEFAULT 'evento',
            criado_por TEXT
        )""",
        f"""CREATE TABLE IF NOT EXISTS solicitacoes (
            id {serial} PRIMARY KEY {autoincrement},
            solicitante TEXT,
            tipo TEXT,
            descricao TEXT,
            data_ref TEXT,
            data_pedido TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            status TEXT DEFAULT 'pendente',
            resposta TEXT,
            respondido_em TIMESTAMP
        )""",
        f"""CREATE TABLE IF NOT EXISTS notificacoes (
            id {serial} PRIMARY KEY {autoincrement},
            destinatario TEXT,
            tipo TEXT,
            mensagem TEXT,
            link TEXT,
            lida INTEGER DEFAULT 0,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        f"""CREATE TABLE IF NOT EXISTS hub_eventos (
            id {serial} PRIMARY KEY {autoincrement},
            titulo TEXT,
            data TEXT,
            inicio TEXT,
            fim TEXT,
            responsavel TEXT,
            criador TEXT,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        f"""CREATE TABLE IF NOT EXISTS equipe_eventos (
            id {serial} PRIMARY KEY {autoincrement},
            titulo TEXT NOT NULL,
            data TEXT NOT NULL,
            local TEXT,
            descricao TEXT,
            criado_por TEXT,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        f"""CREATE TABLE IF NOT EXISTS equipe_eventos_membros (
            id {serial} PRIMARY KEY {autoincrement},
            evento_id INTEGER NOT NULL,
            funcionario TEXT NOT NULL
        )""",
        f"""CREATE TABLE IF NOT EXISTS repositorio (
            id {serial} PRIMARY KEY {autoincrement},
            nome TEXT NOT NULL,
            categoria TEXT NOT NULL,
            drive_id TEXT NOT NULL,
            drive_link TEXT NOT NULL,
            mime_type TEXT,
            tamanho INTEGER,
            autor TEXT,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        f"""CREATE TABLE IF NOT EXISTS atas (
            id {serial} PRIMARY KEY {autoincrement},
            titulo TEXT NOT NULL,
            data_reuniao TEXT,
            presentes TEXT,
            pauta TEXT,
            deliberacoes TEXT,
            link TEXT,
            autor TEXT,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        f"""CREATE TABLE IF NOT EXISTS ramais (
            id {serial} PRIMARY KEY {autoincrement},
            nome TEXT NOT NULL,
            ramal TEXT NOT NULL,
            cargo TEXT,
            setor TEXT DEFAULT 'Geral',
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
    ]

    for sql in tables:
        c.execute(sql)

    # Migração segura para SQLite
    if not is_pg():
        for sql in ["ALTER TABLE solicitacoes ADD COLUMN data_ref TEXT"]:
            try:
                c.execute(sql)
            except:
                pass

    conn.commit()
    conn.close()


# =========================
# 🔐 LOGIN
# =========================
@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        matricula = str(request.form.get("matricula", "")).strip()

        try:
            df = pd.read_csv(FUNCIONARIOS_URL)
            df.columns = df.columns.str.strip()
            df = df.fillna("")
            df["EMAIL INSTITUCIONAL"] = df["EMAIL INSTITUCIONAL"].astype(str).str.strip().str.lower()
            df["MATRÍCULA"] = df["MATRÍCULA"].astype(str).str.strip()

            user = df[
                (df["EMAIL INSTITUCIONAL"] == email) &
                (df["MATRÍCULA"] == matricula)
            ]

            if not user.empty:
                session["nome"] = user.iloc[0]["NOME"]
                session["link_csv_horas"] = user.iloc[0]["LINK_CSV"]
                session["cargo"] = str(user.iloc[0].get("CARGO", "funcionario")).strip().lower() or "funcionario"
                session["planilha_sgrh_url"] = str(user.iloc[0].get("PLANILHA SGRH", "")).strip()

                if session["cargo"] == "chefe":
                    funcionarios = df[["NOME", "LINK_CSV"]].fillna("").to_dict("records")
                    funcionarios = [f for f in funcionarios if f["LINK_CSV"]]
                    session["funcionarios"] = funcionarios

                return redirect("/dashboard")

            return "Login inválido", 401

        except Exception as e:
            print("ERRO LOGIN:", e)
            return "Erro ao acessar planilha", 500

    return render_template("login.html")


# =========================
# 📊 DASHBOARD
# =========================
@app.route("/dashboard")
def dashboard():
    if not session.get("nome"):
        return redirect("/")

    hoje_str = datetime.now().strftime("%Y-%m-%d")
    if session.get("ultima_geracao_notificacoes") != hoje_str:
        conn = get_conn()
        _gerar_notificacoes_automaticas(conn, session["nome"])
        conn.close()
        session["ultima_geracao_notificacoes"] = hoje_str

    return render_template(
        "dashboard.html",
        nome=session["nome"],
        cargo=session.get("cargo", "funcionario")
    )


# =========================
# 🧩 KANBAN
# =========================
@app.route("/tasks")
def get_tasks():
    if not session.get("nome"):
        return jsonify([])
    conn = get_conn()
    c = cursor(conn)
    c.execute("SELECT id, nome, titulo, descricao, status FROM tasks")
    rows = c.fetchall()
    conn.close()
    if is_pg():
        return jsonify([dict(r) for r in rows])
    return jsonify([
        {"id": r[0], "nome": r[1], "titulo": r[2], "descricao": r[3], "status": r[4]}
        for r in rows
    ])


@app.route("/tasks", methods=["POST"])
def create_task():
    if not session.get("nome"):
        return jsonify({"erro": "não logado"}), 401
    data = request.json
    conn = get_conn()
    c = cursor(conn)
    c.execute(
        fix_sql("INSERT INTO tasks (nome, titulo, descricao, status) VALUES (?, ?, ?, ?)"),
        (session["nome"], data.get("titulo"), data.get("descricao", ""), "todo")
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/tasks/<int:id>", methods=["PUT"])
def update_task(id):
    data = request.json
    conn = get_conn()
    c = cursor(conn)
    c.execute(fix_sql("UPDATE tasks SET status=? WHERE id=?"), (data.get("status"), id))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/tasks/<int:id>", methods=["DELETE"])
def delete_task(id):
    if not session.get("nome"):
        return jsonify({"erro": "não logado"}), 401
    conn = get_conn()
    c = cursor(conn)
    c.execute(fix_sql("DELETE FROM tasks WHERE id=?"), (id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


# =========================
# ⏱️ HORAS
# =========================
@app.route("/horas")
def horas():
    if "nome" not in session:
        return redirect("/")

    cargo = session.get("cargo", "funcionario")
    funcionarios = session.get("funcionarios", [])

    if cargo == "chefe":
        nome_selecionado = request.args.get("funcionario")
        if nome_selecionado:
            func = next((f for f in funcionarios if f["NOME"] == nome_selecionado), None)
            url_csv = func["LINK_CSV"] if func else None
        else:
            url_csv = None
    else:
        url_csv = session.get("link_csv_horas")
        nome_selecionado = session["nome"]

    dados_finais = []

    if url_csv:
        try:
            response = requests.get(url_csv)
            if response.status_code == 200:
                df = pd.read_csv(io.StringIO(response.content.decode("utf-8")), sep=None, engine="python")
                df = df.iloc[2:].reset_index(drop=True).iloc[:, :6]
                df.columns = ["DATA", "JUSTIFICATIVA", "PERIODO", "CREDITO", "DEBITO", "SALDO"]
                df = df.fillna("")
                df = df[df["DATA"].str.strip() != ""]
                dados_finais = df.to_dict("records")
        except Exception as e:
            print("ERRO AO LER CSV:", e)

    return render_template(
        "horas.html",
        dados=dados_finais,
        cargo=cargo,
        funcionarios=funcionarios,
        nome_selecionado=nome_selecionado if cargo == "chefe" else session["nome"]
    )


# =========================
# ✏️ LANÇAR HORA
# =========================
def salvar_na_planilha(url_editavel, nova_linha):
    try:
        client = get_gspread_client()
        sh = client.open_by_url(url_editavel).sheet1
        col_data = sh.col_values(1)
        proxima_linha = len(col_data) + 1
        sh.update(
            f"A{proxima_linha}:E{proxima_linha}",
            [nova_linha[:5]],
            value_input_option="USER_ENTERED"
        )
        return True
    except Exception as e:
        print(f"ERRO salvar_na_planilha: {e}")
        return False


@app.route("/lancar-hora", methods=["POST"])
def lancar_hora():
    if "nome" not in session:
        return redirect("/")

    url_csv = session.get("link_csv_horas")
    if not url_csv:
        return "URL da planilha não encontrada na sessão", 400

    url_editavel = extrair_url_editavel(url_csv)
    if not url_editavel:
        return "URL da planilha inválida ou não reconhecida", 400

    nova_linha = [
        request.form.get("data"),
        request.form.get("justificativa"),
        request.form.get("periodo"),
        request.form.get("credito"),
        request.form.get("debito"),
        ""
    ]

    if salvar_na_planilha(url_editavel, nova_linha):
        return redirect("/horas")
    else:
        return "Erro ao salvar na planilha", 500


# =========================
# 📢 COMUNICADOS
# =========================
EMOJIS_REACAO = ["👍", "❤️", "😂", "😮", "😢"]

@app.route("/comunicados", methods=["GET"])
def get_comunicados():
    if not session.get("nome"):
        return jsonify([]), 401

    conn = get_conn()
    c = cursor(conn)
    c.execute("SELECT id, titulo, texto, autor, criado_em FROM comunicados ORDER BY criado_em DESC")
    rows = c.fetchall()

    resultado = []
    for r in rows:
        com_id = r["id"] if is_pg() else r[0]
        titulo = r["titulo"] if is_pg() else r[1]
        texto = r["texto"] if is_pg() else r[2]
        autor = r["autor"] if is_pg() else r[3]
        criado_em = r["criado_em"] if is_pg() else r[4]

        reacoes = {}
        for emoji in EMOJIS_REACAO:
            c.execute(fix_sql("SELECT COUNT(*) FROM reacoes WHERE comunicado_id=? AND emoji=?"), (com_id, emoji))
            row_count = c.fetchone()
            count = row_count[0] if not is_pg() else list(row_count.values())[0]
            c.execute(
                fix_sql("SELECT 1 FROM reacoes WHERE comunicado_id=? AND emoji=? AND usuario=?"),
                (com_id, emoji, session["nome"])
            )
            eu_reagi = c.fetchone() is not None
            reacoes[emoji] = {"count": count, "eu": eu_reagi}

        resultado.append({
            "id": com_id, "titulo": titulo, "texto": texto,
            "autor": autor, "criado_em": str(criado_em), "reacoes": reacoes
        })

    conn.close()
    return jsonify(resultado)


@app.route("/comunicados", methods=["POST"])
def criar_comunicado():
    if not session.get("nome"):
        return jsonify({"erro": "não logado"}), 401
    data = request.json
    conn = get_conn()
    c = cursor(conn)
    c.execute(
        fix_sql("INSERT INTO comunicados (titulo, texto, autor) VALUES (?, ?, ?)"),
        (data["titulo"], data["texto"], session["nome"])
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/comunicados/<int:id>", methods=["DELETE"])
def deletar_comunicado(id):
    if not session.get("nome"):
        return jsonify({"erro": "não logado"}), 401
    conn = get_conn()
    c = cursor(conn)
    c.execute(fix_sql("DELETE FROM comunicados WHERE id=? AND autor=?"), (id, session["nome"]))
    c.execute(fix_sql("DELETE FROM reacoes WHERE comunicado_id=?"), (id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/comunicados/<int:id>/reagir", methods=["POST"])
def reagir_comunicado(id):
    if not session.get("nome"):
        return jsonify({"erro": "não logado"}), 401

    emoji = request.json.get("emoji")
    usuario = session["nome"]

    conn = get_conn()
    c = cursor(conn)

    c.execute(
        fix_sql("SELECT id FROM reacoes WHERE comunicado_id=? AND usuario=? AND emoji=?"),
        (id, usuario, emoji)
    )
    existente = c.fetchone()

    if existente:
        ex_id = existente["id"] if is_pg() else existente[0]
        c.execute(fix_sql("DELETE FROM reacoes WHERE id=?"), (ex_id,))
        acao = "removida"
    else:
        c.execute(
            fix_sql("INSERT INTO reacoes (comunicado_id, usuario, emoji) VALUES (?, ?, ?)"),
            (id, usuario, emoji)
        )
        acao = "adicionada"

    conn.commit()
    conn.close()
    return jsonify({"ok": True, "acao": acao})


# =========================
# 📅 EVENTOS / CALENDÁRIO
# =========================
@app.route("/eventos", methods=["GET"])
def get_eventos():
    if not session.get("nome"):
        return jsonify([]), 401
    conn = get_conn()
    c = cursor(conn)
    c.execute("SELECT id, titulo, data, tipo, criado_por FROM eventos ORDER BY data")
    rows = c.fetchall()
    conn.close()
    if is_pg():
        return jsonify([dict(r) for r in rows])
    return jsonify([
        {"id": r[0], "titulo": r[1], "data": r[2], "tipo": r[3], "criado_por": r[4]}
        for r in rows
    ])


@app.route("/eventos", methods=["POST"])
def criar_evento():
    if not session.get("nome"):
        return jsonify({"erro": "não logado"}), 401
    if session.get("cargo") != "chefe":
        return jsonify({"erro": "sem permissão"}), 403

    data = request.json
    conn = get_conn()
    c = cursor(conn)
    c.execute(
        fix_sql("INSERT INTO eventos (titulo, data, tipo, criado_por) VALUES (?, ?, ?, ?)"),
        (data["titulo"], data["data"], data.get("tipo", "evento"), session["nome"])
    )

    try:
        df = pd.read_csv(FUNCIONARIOS_URL)
        df.columns = df.columns.str.strip()
        df = df.fillna("")
        data_fmt = "/".join(reversed(data["data"].split("-")))
        for _, row in df.iterrows():
            nome_func = str(row.get("NOME", "")).strip()
            if nome_func and nome_func != session["nome"]:
                criar_notificacao(
                    conn, nome_func, "evento",
                    f"📅 Novo evento: {data['titulo']} em {data_fmt}"
                )
    except Exception as e:
        print("ERRO ao notificar evento:", e)

    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/eventos/<int:id>", methods=["DELETE"])
def deletar_evento(id):
    if not session.get("nome") or session.get("cargo") != "chefe":
        return jsonify({"erro": "sem permissão"}), 403
    conn = get_conn()
    c = cursor(conn)
    c.execute(fix_sql("DELETE FROM eventos WHERE id=?"), (id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


# =========================
# 🎂 ANIVERSARIANTES
# =========================
@app.route("/aniversariantes")
def aniversariantes():
    if not session.get("nome"):
        return jsonify([]), 401
    try:
        df = pd.read_csv(FUNCIONARIOS_URL)
        df.columns = df.columns.str.strip()
        df = df.fillna("")
        hoje = pd.Timestamp.now()
        mes_atual = hoje.month
        result = []
        for _, row in df.iterrows():
            nasc = str(row.get("NASCIMENTO", "")).strip()
            if not nasc:
                continue
            try:
                partes = nasc.split("/")
                dia, mes = int(partes[0]), int(partes[1])
                result.append({
                    "nome": row["NOME"],
                    "dia": dia,
                    "mes": mes,
                    "hoje": dia == hoje.day and mes == mes_atual
                })
            except:
                continue
        result = [r for r in result if r["mes"] == mes_atual]
        result.sort(key=lambda x: x["dia"])
        return jsonify(result)
    except Exception as e:
        print("ERRO ANIVERSARIANTES:", e)
        return jsonify([])


# =========================
# 📋 SOLICITAÇÕES
# =========================
@app.route("/solicitacoes", methods=["GET"])
def get_solicitacoes():
    if not session.get("nome"):
        return jsonify([]), 401

    conn = get_conn()
    c = cursor(conn)

    if session.get("cargo") == "chefe":
        c.execute("""
            SELECT id, solicitante, tipo, descricao, data_ref, data_pedido, status, resposta, respondido_em
            FROM solicitacoes ORDER BY data_pedido DESC
        """)
    else:
        c.execute(
            fix_sql("""
                SELECT id, solicitante, tipo, descricao, data_ref, data_pedido, status, resposta, respondido_em
                FROM solicitacoes WHERE solicitante=? ORDER BY data_pedido DESC
            """),
            (session["nome"],)
        )

    rows = c.fetchall()
    conn.close()

    if is_pg():
        return jsonify([dict(r) for r in rows])
    return jsonify([{
        "id": r[0], "solicitante": r[1], "tipo": r[2],
        "descricao": r[3], "data_ref": r[4], "data_pedido": r[5],
        "status": r[6], "resposta": r[7], "respondido_em": r[8]
    } for r in rows])


@app.route("/solicitacoes", methods=["POST"])
def criar_solicitacao():
    if not session.get("nome"):
        return jsonify({"erro": "não logado"}), 401

    data = request.json
    nome = session["nome"]

    conn = get_conn()
    c = cursor(conn)
    c.execute(
        fix_sql("""
            INSERT INTO solicitacoes (solicitante, tipo, descricao, data_ref, status)
            VALUES (?, ?, ?, ?, 'pendente')
        """),
        (nome, data.get("tipo"), data.get("descricao", ""), data.get("data_ref", ""))
    )

    try:
        df = pd.read_csv(FUNCIONARIOS_URL)
        df.columns = df.columns.str.strip()
        df = df.fillna("")
        chefes = df[df["CARGO"].str.strip().str.lower() == "chefe"]["NOME"].tolist()
        for chefe in chefes:
            criar_notificacao(
                conn, chefe, "solicitacao",
                f"📋 Nova solicitação de {nome}: {data.get('tipo')}"
            )
    except Exception as e:
        print("ERRO ao notificar chefe:", e)

    conn.commit()
    try:
        df_email = pd.read_csv(FUNCIONARIOS_URL)
        df_email.columns = df_email.columns.str.strip()
        df_email = df_email.fillna("")
        df_email["NOME"] = df_email["NOME"].astype(str).str.strip()
        row_email = df_email[df_email["NOME"] == nome]
        email_dest = str(row_email.iloc[0].get("EMAIL INSTITUCIONAL", "")).strip() if not row_email.empty else ""
        if email_dest:
            email_solicitacao_confirmacao(email_dest, nome, data.get("tipo"), data.get("descricao", ""))
    except Exception as e:
        print(f"⚠️ Erro ao enviar email de confirmação: {e}")
    conn.close()
    return jsonify({"ok": True})


@app.route("/solicitacoes/<int:id>/responder", methods=["POST"])
def responder_solicitacao(id):
    if not session.get("nome") or session.get("cargo") != "chefe":
        return jsonify({"erro": "sem permissão"}), 403

    data = request.json
    conn = get_conn()
    c = cursor(conn)

    c.execute(fix_sql("SELECT solicitante, tipo FROM solicitacoes WHERE id=?"), (id,))
    row = c.fetchone()

    c.execute(
        fix_sql("""
            UPDATE solicitacoes
            SET status=?, resposta=?, respondido_em=CURRENT_TIMESTAMP
            WHERE id=?
        """),
        (data["status"], data.get("resposta", ""), id)
    )

    if row:
        solicitante = row["solicitante"] if is_pg() else row[0]
        tipo = row["tipo"] if is_pg() else row[1]
        emoji = "✅" if data["status"] == "aprovada" else "❌"
        criar_notificacao(
            conn, solicitante, "resposta",
            f"{emoji} Sua solicitação de '{tipo}' foi {data['status']}."
        )

    conn.commit()
    try:
        df_email = pd.read_csv(FUNCIONARIOS_URL)
        df_email.columns = df_email.columns.str.strip()
        df_email = df_email.fillna("")
        df_email["NOME"] = df_email["NOME"].astype(str).str.strip()
        row_email = df_email[df_email["NOME"] == solicitante]
        email_dest = str(row_email.iloc[0].get("EMAIL INSTITUCIONAL", "")).strip() if not row_email.empty else ""
        if email_dest:
            email_solicitacao_resposta(email_dest, solicitante, tipo, data["status"], data.get("resposta", ""))
    except Exception as e:
        print(f"⚠️ Erro ao enviar email de resposta: {e}")
    conn.close()
    return jsonify({"ok": True})


@app.route("/solicitacoes/<int:id>", methods=["DELETE"])
def deletar_solicitacao(id):
    if not session.get("nome"):
        return jsonify({"erro": "não logado"}), 401
    if session.get("cargo") != "chefe":
        return jsonify({"erro": "sem permissão"}), 403

    conn = get_conn()
    c = cursor(conn)
    c.execute(fix_sql("DELETE FROM solicitacoes WHERE id=?"), (id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


# =========================
# 🔔 NOTIFICAÇÕES
# =========================
@app.route("/notificacoes")
def get_notificacoes():
    if not session.get("nome"):
        return jsonify([]), 401

    conn = get_conn()
    c = cursor(conn)
    c.execute(
        fix_sql("""
            SELECT id, tipo, mensagem, link, lida, criado_em
            FROM notificacoes
            WHERE destinatario=?
            ORDER BY criado_em DESC
            LIMIT 30
        """),
        (session["nome"],)
    )
    rows = c.fetchall()
    conn.close()

    if is_pg():
        return jsonify([dict(r) for r in rows])
    return jsonify([{
        "id": r[0], "tipo": r[1], "mensagem": r[2],
        "link": r[3], "lida": r[4], "criado_em": r[5]
    } for r in rows])


@app.route("/notificacoes/marcar-lidas", methods=["POST"])
def marcar_notificacoes_lidas():
    if not session.get("nome"):
        return jsonify({"erro": "não logado"}), 401

    ids = request.json.get("ids", [])
    conn = get_conn()
    c = cursor(conn)

    if ids:
        if is_pg():
            c.execute(
                "UPDATE notificacoes SET lida=1 WHERE id = ANY(%s) AND destinatario=%s",
                (ids, session["nome"])
            )
        else:
            placeholders = ",".join("?" * len(ids))
            c.execute(
                f"UPDATE notificacoes SET lida=1 WHERE id IN ({placeholders}) AND destinatario=?",
                ids + [session["nome"]]
            )
    else:
        c.execute(fix_sql("UPDATE notificacoes SET lida=1 WHERE destinatario=?"), (session["nome"],))

    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/notificacoes/<int:id>", methods=["DELETE"])
def deletar_notificacao(id):
    if not session.get("nome"):
        return jsonify({"erro": "não logado"}), 401
    conn = get_conn()
    c = cursor(conn)
    c.execute(fix_sql("DELETE FROM notificacoes WHERE id=? AND destinatario=?"), (id, session["nome"]))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


def _gerar_notificacoes_automaticas(conn, nome_usuario):
    c = cursor(conn)
    hoje = datetime.now()
    amanha = hoje + timedelta(days=1)
    em_3_dias = hoje + timedelta(days=3)
    hoje_str = hoje.strftime("%Y-%m-%d")

    try:
        c.execute(
            fix_sql("SELECT id, titulo, data FROM eventos WHERE data >= ? AND data <= ?"),
            (hoje_str, em_3_dias.strftime("%Y-%m-%d"))
        )
        eventos = c.fetchall()
        for ev in eventos:
            ev_id = ev["id"] if is_pg() else ev[0]
            titulo = ev["titulo"] if is_pg() else ev[1]
            data_str = ev["data"] if is_pg() else ev[2]

            c.execute(
                fix_sql("""
                    SELECT 1 FROM notificacoes
                    WHERE destinatario=? AND tipo='evento_proximo'
                    AND mensagem LIKE ? AND DATE(criado_em)=?
                """),
                (nome_usuario, f"%{titulo}%", hoje_str)
            )
            if not c.fetchone():
                data_fmt = "/".join(reversed(str(data_str).split("-")))
                dias_faltam = (datetime.strptime(str(data_str), "%Y-%m-%d") - hoje).days
                if dias_faltam == 0:
                    msg = f"⏰ Hoje tem: '{titulo}'!"
                elif dias_faltam == 1:
                    msg = f"⏰ Amanhã tem: '{titulo}'!"
                else:
                    msg = f"⏰ Em {dias_faltam} dias: '{titulo}' ({data_fmt})"
                criar_notificacao(conn, nome_usuario, "evento_proximo", msg)
    except Exception as e:
        print("ERRO notif evento:", e)

    try:
        df = pd.read_csv(FUNCIONARIOS_URL)
        df.columns = df.columns.str.strip()
        df = df.fillna("")
        for _, row in df.iterrows():
            nasc = str(row.get("NASCIMENTO", "")).strip()
            nome_aniv = str(row.get("NOME", "")).strip()
            if not nasc or not nome_aniv:
                continue
            try:
                partes = nasc.split("/")
                dia, mes = int(partes[0]), int(partes[1])

                if dia == hoje.day and mes == hoje.month:
                    c.execute(
                        fix_sql("""
                            SELECT 1 FROM notificacoes
                            WHERE destinatario=? AND tipo='aniversario'
                            AND mensagem LIKE ? AND DATE(criado_em)=?
                        """),
                        (nome_usuario, f"%{nome_aniv}%hoje%", hoje_str)
                    )
                    if not c.fetchone():
                        criar_notificacao(
                            conn, nome_usuario, "aniversario",
                            f"🎂 Hoje é aniversário de {nome_aniv}! Não esqueça de parabenizar 🎉"
                        )

                elif dia == amanha.day and mes == amanha.month:
                    c.execute(
                        fix_sql("""
                            SELECT 1 FROM notificacoes
                            WHERE destinatario=? AND tipo='aniversario'
                            AND mensagem LIKE ? AND DATE(criado_em)=?
                        """),
                        (nome_usuario, f"%{nome_aniv}%amanhã%", hoje_str)
                    )
                    if not c.fetchone():
                        criar_notificacao(
                            conn, nome_usuario, "aniversario",
                            f"🎂 Amanhã é aniversário de {nome_aniv}!"
                        )
            except:
                continue
    except Exception as e:
        print("ERRO notif aniversario:", e)

    conn.commit()

# =========================
# 📄 SGRH
# =========================
@app.route("/sgrh")
def sgrh():
    if not session.get("nome"):
        return redirect("/")
    cargo = session.get("cargo", "funcionario")
    funcionarios = session.get("funcionarios", [])
    return render_template("sgrh.html", nome=session["nome"], cargo=cargo, funcionarios=funcionarios)


@app.route("/sgrh/dados")
def sgrh_dados():
    if not session.get("nome"):
        return jsonify({"erro": "não logado"}), 401
    nome_func = request.args.get("funcionario", "").strip()
    if session.get("cargo") != "chefe":
        nome_func = session["nome"]
    url_sgrh = get_url_sgrh_funcionario(nome_func)
    if not url_sgrh:
        return jsonify({"erro": f"URL SGRH não encontrada para '{nome_func}'."})
    try:
        sheet_id, gid = extrair_sheet_e_gid(url_sgrh)
        ws = abrir_aba_por_gid(sheet_id, gid)
        return jsonify({"ok": True, "q1": ler_quinzena(ws, LINHA_Q1), "q2": ler_quinzena(ws, LINHA_Q2)})
    except Exception as e:
        print("ERRO sgrh_dados:", e)
        return jsonify({"erro": str(e)})


@app.route("/sgrh/salvar", methods=["POST"])
def sgrh_salvar():
    if not session.get("nome"):
        return jsonify({"erro": "não logado"}), 401
    data = request.json
    nome_func = data.get("funcionario", "").strip()
    if session.get("cargo") != "chefe":
        nome_func = session["nome"]
    url_sgrh = get_url_sgrh_funcionario(nome_func)
    if not url_sgrh:
        return jsonify({"erro": f"URL SGRH não encontrada para '{nome_func}'."})
    try:
        sheet_id, gid = extrair_sheet_e_gid(url_sgrh)
        ws = abrir_aba_por_gid(sheet_id, gid)
        salvar_quinzena(ws, LINHA_Q1, data.get("q1", []))
        salvar_quinzena(ws, LINHA_Q2, data.get("q2", []))
        return jsonify({"ok": True})
    except Exception as e:
        print("ERRO sgrh_salvar:", e)
        return jsonify({"erro": str(e)})

@app.route("/equipe")
def equipe():
    if not session.get("nome"):
        return redirect("/")
    return render_template(
        "equipe.html",
        nome=session["nome"],
        cargo=session.get("cargo", "funcionario")
    )

@app.route("/equipe/editar-status", methods=["POST"])
def equipe_editar_status():
    if not session.get("nome"):
        return jsonify({"erro": "não logado"}), 401
    if session.get("cargo", "").strip().lower() != "chefe":
        return jsonify({"erro": "sem permissão"}), 403

    data = request.json
    nome_func   = str(data.get("nome", "")).strip()
    novo_status = str(data.get("status", "")).strip().upper()
    novo_cargo  = str(data.get("cargo", "")).strip()

    if not nome_func:
        return jsonify({"erro": "Nome não informado"}), 400

    SHEET_ID = "1nYgCV6SgPf5PTIIJZLQcM9xfw5L08SNber7ZEDfUWOo"

    try:
        client = get_gspread_client()
        sh = client.open_by_key(SHEET_ID)
        ws = sh.worksheet("FUNCIONARIOS")

        col_nomes = ws.col_values(2)
        try:
            linha = col_nomes.index(nome_func) + 1
        except ValueError:
            return jsonify({"erro": f"Funcionário '{nome_func}' não encontrado na planilha"}), 404

        ws.update_cell(linha, 5, novo_status)
        if novo_cargo:
            ws.update_cell(linha, 8, novo_cargo)

        return jsonify({"ok": True})

    except Exception as e:
        print("ERRO equipe_editar_status:", e)
        return jsonify({"erro": str(e)}), 500

@app.route("/notificacoes/limpar-todas", methods=["DELETE"])
def limpar_todas_notificacoes():
    if not session.get("nome"):
        return jsonify({"erro": "não logado"}), 401
    conn = get_conn()
    c = cursor(conn)
    c.execute(fix_sql("DELETE FROM notificacoes WHERE destinatario=?"), (session["nome"],))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/ci")
def ci():
    if not session.get("nome"):
        return redirect("/")
    return render_template(
        "ci.html",
        nome=session["nome"],
        cargo=session.get("cargo", "funcionario"),
        email=session.get("email", ""),
        funcionarios=session.get("funcionarios", [])
    )

@app.route("/ci/enviar", methods=["POST"])
def ci_enviar():
    if not session.get("nome"):
        return jsonify({"erro": "não logado"}), 401

    data   = request.json
    tipo   = str(data.get("tipo", "")).strip()
    datas  = str(data.get("data", "")).strip()
    obs    = str(data.get("obs", "")).strip()

    if not tipo or not datas:
        return jsonify({"erro": "Campos obrigatórios faltando"}), 400

    nome_func = session["nome"]

    try:
        df = pd.read_csv(FUNCIONARIOS_URL)
        df.columns = df.columns.str.strip()
        df = df.fillna("")
        df["NOME"] = df["NOME"].astype(str).str.strip()
        row = df[df["NOME"] == nome_func]
        email_func = str(row.iloc[0].get("EMAIL INSTITUCIONAL", "")).strip() if not row.empty else ""
    except Exception as e:
        print("ERRO ao buscar email:", e)
        email_func = ""

    carimbo = datetime.now().strftime("%m/%d/%Y %H:%M:%S")
    nova_linha = [carimbo, email_func, nome_func, tipo, datas, obs, ""]

    try:
        url_csv = session.get("link_csv_horas", "")
        if not url_csv:
            return jsonify({"erro": "Planilha do funcionário não encontrada na sessão"}), 400

        sheet_match = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url_csv)
        if not sheet_match:
            return jsonify({"erro": "URL de planilha inválida"}), 400
        sheet_id = sheet_match.group(1)

        client = get_gspread_client()
        sh     = client.open_by_key(sheet_id)

        try:
            ws = sh.worksheet("CI")
        except Exception:
            ws = sh.add_worksheet(title="CI", rows=200, cols=10)
            ws.append_row(
                ["Carimbo de data/hora", "Endereço de e-mail", "Funcionário:",
                 "Solicitação / Comunicação:", "Data:", "Observações:", "Devolutiva"],
                value_input_option="USER_ENTERED"
            )

        ws.append_row(nova_linha, value_input_option="USER_ENTERED")
        if email_func:
            email_ci_confirmacao(email_func, nome_func, tipo, datas, obs)
        return jsonify({"ok": True})

    except Exception as e:
        print("ERRO ci_enviar:", e)
        return jsonify({"erro": str(e)}), 500


@app.route("/ci/historico")
def ci_historico():
    if not session.get("nome"):
        return jsonify([]), 401

    resultado = []

    try:
        if session.get("cargo") == "chefe":
            nome_func = request.args.get("funcionario", "").strip()
            if not nome_func:
                return jsonify([])

            df_func = pd.read_csv(FUNCIONARIOS_URL)
            df_func.columns = df_func.columns.str.strip()
            df_func = df_func.fillna("")
            df_func["NOME"] = df_func["NOME"].astype(str).str.strip()

            row_func = df_func[df_func["NOME"] == nome_func]
            if row_func.empty:
                return jsonify({"erro": f"Funcionário '{nome_func}' não encontrado"}), 404

            url_csv = str(row_func.iloc[0].get("LINK_CSV", "")).strip()
            if not url_csv:
                return jsonify([])

            sheet_match = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url_csv)
            if not sheet_match:
                return jsonify([])
            sheet_id = sheet_match.group(1)

            try:
                client = get_gspread_client()
                sh = client.open_by_key(sheet_id)
                ws = sh.worksheet("CI")
                rows = ws.get_all_values()
                for r in rows[1:]:
                    if len(r) < 4 or not r[0]:
                        continue
                    resultado.append({
                        "carimbo":     r[0] if len(r) > 0 else "",
                        "email":       r[1] if len(r) > 1 else "",
                        "funcionario": r[2] if len(r) > 2 else nome_func,
                        "tipo":        r[3] if len(r) > 3 else "",
                        "data":        r[4] if len(r) > 4 else "",
                        "obs":         r[5] if len(r) > 5 else "",
                        "devolutiva":  r[6] if len(r) > 6 else "",
                    })
            except Exception as e:
                print(f"ERRO lendo CI de {nome_func}: {e}")
                return jsonify({"erro": str(e)}), 500

        else:
            url_csv = session.get("link_csv_horas", "")
            if not url_csv:
                return jsonify([])

            sheet_match = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url_csv)
            if not sheet_match:
                return jsonify([])
            sheet_id = sheet_match.group(1)

            try:
                client = get_gspread_client()
                sh = client.open_by_key(sheet_id)
                ws = sh.worksheet("CI")
                rows = ws.get_all_values()
                for r in rows[1:]:
                    if len(r) < 4 or not r[0]:
                        continue
                    resultado.append({
                        "carimbo":     r[0] if len(r) > 0 else "",
                        "email":       r[1] if len(r) > 1 else "",
                        "funcionario": r[2] if len(r) > 2 else session["nome"],
                        "tipo":        r[3] if len(r) > 3 else "",
                        "data":        r[4] if len(r) > 4 else "",
                        "obs":         r[5] if len(r) > 5 else "",
                        "devolutiva":  r[6] if len(r) > 6 else "",
                    })
            except Exception:
                return jsonify([])

        resultado.sort(key=lambda x: x["carimbo"], reverse=True)
        return jsonify(resultado)

    except Exception as e:
        print("ERRO ci_historico:", e)
        return jsonify([])

@app.route("/ci/devolutiva", methods=["POST"])
def ci_devolutiva():
    if not session.get("nome"):
        return jsonify({"erro": "não logado"}), 401
    if session.get("cargo") != "chefe":
        return jsonify({"erro": "sem permissão"}), 403

    data          = request.json
    nome_func     = str(data.get("funcionario", "")).strip()
    carimbo_busca = str(data.get("carimbo", "")).strip()
    devolutiva    = str(data.get("devolutiva", "")).strip()

    if not nome_func or not carimbo_busca or not devolutiva:
        return jsonify({"erro": "Parâmetros insuficientes"}), 400

    try:
        df_func = pd.read_csv(FUNCIONARIOS_URL)
        df_func.columns = df_func.columns.str.strip()
        df_func = df_func.fillna("")
        df_func["NOME"] = df_func["NOME"].astype(str).str.strip()

        row_func = df_func[df_func["NOME"] == nome_func]
        if row_func.empty:
            return jsonify({"erro": f"Funcionário '{nome_func}' não encontrado"}), 404

        url_csv = str(row_func.iloc[0].get("LINK_CSV", "")).strip()
        if not url_csv:
            return jsonify({"erro": "Planilha do funcionário não encontrada"}), 400

        sheet_match = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url_csv)
        if not sheet_match:
            return jsonify({"erro": "URL inválida"}), 400
        sheet_id = sheet_match.group(1)

        client = get_gspread_client()
        sh = client.open_by_key(sheet_id)

        try:
            ws = sh.worksheet("CI")
        except Exception:
            return jsonify({"erro": "Aba CI não encontrada na planilha do funcionário"}), 404

        col_carimbo = ws.col_values(1)
        try:
            linha = col_carimbo.index(carimbo_busca) + 1
        except ValueError:
            return jsonify({"erro": "Linha não encontrada pelo carimbo"}), 404

        ws.update_cell(linha, 7, devolutiva)

        try:
            linha_dados = ws.row_values(linha)
            email_dest  = linha_dados[1] if len(linha_dados) > 1 else ""
            tipo_ci     = linha_dados[3] if len(linha_dados) > 3 else ""
            datas_ci    = linha_dados[4] if len(linha_dados) > 4 else ""
            if email_dest:
                email_ci_devolutiva(email_dest, nome_func, tipo_ci, datas_ci, devolutiva)
        except Exception as e:
            print(f"⚠️ Erro ao enviar email de devolutiva: {e}")

        conn = get_conn()
        emoji_map = {"Deferido": "✅", "Indeferido": "❌", "Ciente": "🔵"}
        emoji = emoji_map.get(devolutiva, "📋")
        criar_notificacao(
            conn, nome_func, "ci",
            f"{emoji} Sua CI foi marcada como '{devolutiva}' pela chefia."
        )
        conn.commit()
        conn.close()

        return jsonify({"ok": True})

    except Exception as e:
        print("ERRO ci_devolutiva:", e)
        return jsonify({"erro": str(e)}), 500

# =========================
# 📁 REPOSITÓRIO (Google Drive)
# =========================
REPOSITORIO_FOLDER_ID = "1fi4DRiN7ei1x6TEipQJjONvkoVWCbc9t"


def get_drive_service():
    scope = [
        "https://www.googleapis.com/auth/drive",
        "https://www.googleapis.com/auth/spreadsheets",
    ]
    cred_json = os.environ.get("GOOGLE_CREDENTIALS")
    cred_dict = json.loads(cred_json)
    creds = Credentials.from_service_account_info(cred_dict, scopes=scope)
    return build("drive", "v3", credentials=creds)


@app.route("/repositorio")
def repositorio():
    if not session.get("nome"):
        return redirect("/")
    return render_template(
        "repositorio.html",
        nome=session["nome"],
        cargo=session.get("cargo", "funcionario")
    )


@app.route("/repositorio/dados")
def repositorio_dados():
    if not session.get("nome"):
        return jsonify([]), 401
    conn = get_conn()
    c = cursor(conn)
    c.execute("""
        SELECT id, nome, categoria, drive_link, mime_type, tamanho, autor, criado_em
        FROM repositorio
        ORDER BY criado_em DESC
    """)
    rows = c.fetchall()
    conn.close()
    if is_pg():
        return jsonify([dict(r) for r in rows])
    return jsonify([{
        "id": r[0], "nome": r[1], "categoria": r[2], "link": r[3],
        "mime_type": r[4], "tamanho": r[5], "autor": r[6], "criado_em": r[7],
    } for r in rows])


@app.route("/repositorio", methods=["POST"])
def repositorio_upload():
    if not session.get("nome"):
        return jsonify({"erro": "não logado"}), 401

    arquivo   = request.files.get("arquivo")
    nome      = request.form.get("nome", "").strip()
    categoria = request.form.get("categoria", "Outros").strip()

    if not arquivo or not nome:
        return jsonify({"erro": "Arquivo e nome são obrigatórios"}), 400

    conteudo  = arquivo.read()
    mime_type = arquivo.content_type or "application/octet-stream"
    tamanho   = len(conteudo)

    try:
        service = get_drive_service()
        file_metadata = {
            "name":    arquivo.filename,
            "parents": [REPOSITORIO_FOLDER_ID],
        }
        media = MediaIoBaseUpload(
            io.BytesIO(conteudo),
            mimetype=mime_type,
            resumable=False,
        )
        uploaded = service.files().create(
            body=file_metadata,
            media_body=media,
            fields="id, webViewLink",
            supportsAllDrives=True,
        ).execute()

        drive_id   = uploaded.get("id")
        drive_link = uploaded.get("webViewLink", f"https://drive.google.com/file/d/{drive_id}/view")

        service.permissions().create(
            fileId=drive_id,
            body={"type": "anyone", "role": "reader"},
            supportsAllDrives=True,
        ).execute()

    except Exception as e:
        print("ERRO upload Drive:", e)
        return jsonify({"erro": f"Erro ao enviar para o Drive: {str(e)}"}), 500

    conn = get_conn()
    c = cursor(conn)
    c.execute(
        fix_sql("""
            INSERT INTO repositorio (nome, categoria, drive_id, drive_link, mime_type, tamanho, autor)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """),
        (nome, categoria, drive_id, drive_link, mime_type, tamanho, session["nome"])
    )
    conn.commit()
    conn.close()

    return jsonify({"ok": True, "link": drive_link})


@app.route("/repositorio/<int:id>", methods=["DELETE"])
def repositorio_deletar(id):
    if not session.get("nome"):
        return jsonify({"erro": "não logado"}), 401

    conn = get_conn()
    c = cursor(conn)

    c.execute(fix_sql("SELECT drive_id, autor FROM repositorio WHERE id=?"), (id,))
    row = c.fetchone()

    if not row:
        conn.close()
        return jsonify({"erro": "Documento não encontrado"}), 404

    drive_id = row["drive_id"] if is_pg() else row[0]
    autor    = row["autor"] if is_pg() else row[1]

    if autor != session["nome"] and session.get("cargo") != "chefe":
        conn.close()
        return jsonify({"erro": "Sem permissão"}), 403

    try:
        service = get_drive_service()
        service.files().delete(fileId=drive_id).execute()
    except Exception as e:
        print(f"AVISO: erro ao deletar do Drive: {e}")

    c.execute(fix_sql("DELETE FROM repositorio WHERE id=?"), (id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/avisos")
def avisos():
    return render_template("avisos.html", nome=session['nome'])

@app.route("/ocorrencias")
def solicitacoes_page():
    if not session.get("nome"):
        return redirect("/")
    return render_template(
        "ocorrencias.html",
        nome=session["nome"],
        cargo=session.get("cargo", "funcionario")
    )

# =========================
# 🗓️ CALENDÁRIO LETIVO
# =========================
CALENDARIO_LETIVO_SHEET_ID = "1lpi2pWcFBMw4bwCSkzcXw8D__7UPQcH49QUN9L7lvwM"
CALENDARIO_LETIVO_ABA = "CALENDÁRIO LETIVO"

@app.route("/calendario_letivo")
def calendario_letivo():
    if not session.get("nome"):
        return redirect("/")
    return render_template(
        "calendario_letivo.html",
        nome=session["nome"],
        cargo=session.get("cargo", "funcionario")
    )


@app.route("/calendario_letivo/dados")
def calendario_letivo_dados():
    if not session.get("nome"):
        return jsonify({"erro": "não logado"}), 401

    try:
        client   = get_gspread_client()
        sh       = client.open_by_key(CALENDARIO_LETIVO_SHEET_ID)
        ws       = sh.worksheet(CALENDARIO_LETIVO_ABA)
        valores  = ws.get_all_values()

        if not valores:
            return jsonify({"erro": "Planilha vazia"}), 400

        MESES_MAP = {
            "JAN": 1,  "FEV": 2,  "MAR": 3,  "ABR": 4,
            "MAI": 5,  "JUN": 6,  "JUL": 7,  "AGO": 8,
            "SET": 9,  "OUT": 10, "NOV": 11, "DEZ": 12,
        }

        meses_resultado = []

        for linha in valores[1:]:
            if not linha:
                continue
            nome_mes = str(linha[0]).strip().upper()
            num_mes  = MESES_MAP.get(nome_mes)
            if not num_mes:
                continue

            dias = {}
            for col_idx in range(1, 32):
                dia_num = col_idx
                if col_idx < len(linha):
                    valor = str(linha[col_idx]).strip()
                    if valor:
                        dias[str(dia_num)] = valor

            meses_resultado.append({"mes": num_mes, "dias": dias})

        meses_resultado.sort(key=lambda x: x["mes"])

        return jsonify({"meses": meses_resultado})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"erro": str(e) or "Erro desconhecido"}), 500

HUB_SHEET_ID = "1IBut2tYJV1g7gC4c0jMHGbB1cqVB9ic0xwWlc8Tx0vk"

@app.route("/hub")
def hub():
    if not session.get("nome"):
        return redirect("/")
    return render_template("hub.html", nome=session['nome'], cargo=session.get("cargo", "funcionario"))

@app.route("/hub/dados")
def hub_dados():
    if not session.get("nome"):
        return jsonify({"erro": "não logado"}), 401
    try:
        client = get_gspread_client()
        ws = client.open_by_key(HUB_SHEET_ID).worksheet("HUB")
        rows = ws.get_all_values()
        eventos = []
        for i, r in enumerate(rows[1:], start=2):
            if not r[0]:
                continue
            eventos.append({
                "id":          r[0],
                "data":        r[1] if len(r) > 1 else "",
                "titulo":      r[2] if len(r) > 2 else "",
                "inicio":      r[3] if len(r) > 3 else "",
                "fim":         r[4] if len(r) > 4 else "",
                "responsavel": r[5] if len(r) > 5 else "",
                "status":      "aprovado",
                "criador":     "",
                "linha":       i
            })
        return jsonify({"eventos": eventos})
    except Exception as e:
        return jsonify({"erro": str(e)}), 500

@app.route("/hub/criar", methods=["POST"])
def hub_criar():
    if not session.get("nome"):
        return jsonify({"erro": "não logado"}), 401
    try:
        d = request.json
        client = get_gspread_client()
        ws = client.open_by_key(HUB_SHEET_ID).worksheet("HUB")

        rows = ws.get_all_values()
        ids = [int(r[0]) for r in rows[1:] if r[0].isdigit()]
        novo_id = max(ids) + 1 if ids else 1

        ws.append_row(
            [novo_id, d.get("data",""), d.get("titulo",""), d.get("inicio",""), d.get("fim",""), d.get("responsavel",""), "aprovado"],
            value_input_option="USER_ENTERED"
        )
        return jsonify({"ok": True})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"erro": str(e)}), 500

@app.route("/hub/excluir", methods=["POST"])
def hub_excluir():
    if not session.get("nome"):
        return jsonify({"erro": "não logado"}), 401
    try:
        d = request.json
        evento_id = str(d.get("id"))
        client = get_gspread_client()
        ws = client.open_by_key(HUB_SHEET_ID).worksheet("HUB")
        col_ids = ws.col_values(1)
        try:
            linha = col_ids.index(evento_id) + 1
            ws.delete_rows(linha)
        except ValueError:
            return jsonify({"erro": "Evento não encontrado"}), 404
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"erro": str(e)}), 500

# =========================
# 🎯 EQUIPES / EVENTOS
# =========================
@app.route("/evento")
def evento():
    if not session.get("nome"):
        return redirect("/")
    return render_template("evento.html", nome=session['nome'], cargo=session.get("cargo", "funcionario"))


@app.route("/equipes-eventos/listar", methods=["GET"])
def equipes_eventos_listar():
    if not session.get("nome"):
        return jsonify([]), 401
    conn = get_conn()
    c = cursor(conn)
    c.execute("""
        SELECT e.id, e.titulo, e.data, e.local, e.descricao, e.criado_por, e.criado_em
        FROM equipe_eventos e
        ORDER BY e.data DESC
    """)
    eventos = c.fetchall()

    resultado = []
    for ev in eventos:
        ev_id    = ev["id"] if is_pg() else ev[0]
        titulo   = ev["titulo"] if is_pg() else ev[1]
        data_ev  = ev["data"] if is_pg() else ev[2]
        local_ev = ev["local"] if is_pg() else ev[3]
        desc     = ev["descricao"] if is_pg() else ev[4]
        criado_por = ev["criado_por"] if is_pg() else ev[5]
        criado_em  = ev["criado_em"] if is_pg() else ev[6]

        c.execute(fix_sql("SELECT funcionario FROM equipe_eventos_membros WHERE evento_id = ?"), (ev_id,))
        membros_rows = c.fetchall()
        membros = [r["funcionario"] if is_pg() else r[0] for r in membros_rows]

        resultado.append({
            "id": ev_id, "titulo": titulo, "data": data_ev,
            "local": local_ev, "descricao": desc,
            "criado_por": criado_por, "criado_em": str(criado_em),
            "equipe": membros
        })

    conn.close()
    return jsonify(resultado)


@app.route("/equipes-eventos/criar", methods=["POST"])
def equipes_eventos_criar():
    if not session.get("nome"):
        return jsonify({"erro": "não logado"}), 401

    data = request.json
    titulo    = str(data.get("titulo", "")).strip()
    data_ev   = str(data.get("data", "")).strip()
    local_ev  = str(data.get("local", "")).strip()
    descricao = str(data.get("descricao", "")).strip()
    equipe    = data.get("equipe", [])

    if not titulo or not data_ev:
        return jsonify({"erro": "Título e data são obrigatórios"}), 400

    conn = get_conn()
    c = cursor(conn)

    if is_pg():
        c.execute("""
            INSERT INTO equipe_eventos (titulo, data, local, descricao, criado_por)
            VALUES (%s, %s, %s, %s, %s) RETURNING id
        """, (titulo, data_ev, local_ev, descricao, session["nome"]))
        ev_id = c.fetchone()["id"]
    else:
        c.execute("""
            INSERT INTO equipe_eventos (titulo, data, local, descricao, criado_por)
            VALUES (?, ?, ?, ?, ?)
        """, (titulo, data_ev, local_ev, descricao, session["nome"]))
        ev_id = c.lastrowid

    for membro in equipe:
        membro = str(membro).strip()
        if membro:
            c.execute(fix_sql("""
                INSERT INTO equipe_eventos_membros (evento_id, funcionario)
                VALUES (?, ?)
            """), (ev_id, membro))

    data_fmt = "/".join(reversed(data_ev.split("-"))) if "-" in data_ev else data_ev
    for membro in equipe:
        membro = str(membro).strip()
        if membro and membro != session["nome"]:
            criar_notificacao(
                conn, membro, "evento",
                f"🎯 Você foi adicionado à equipe do evento: {titulo} ({data_fmt})"
            )

    conn.commit()
    conn.close()
    return jsonify({"ok": True, "id": ev_id})


@app.route("/equipes-eventos/<int:id>", methods=["PUT"])
def equipes_eventos_editar(id):
    if not session.get("nome"):
        return jsonify({"erro": "não logado"}), 401

    data = request.json
    titulo    = str(data.get("titulo", "")).strip()
    data_ev   = str(data.get("data", "")).strip()
    local_ev  = str(data.get("local", "")).strip()
    descricao = str(data.get("descricao", "")).strip()
    equipe    = data.get("equipe", [])

    if not titulo or not data_ev:
        return jsonify({"erro": "Título e data são obrigatórios"}), 400

    conn = get_conn()
    c = cursor(conn)

    c.execute(fix_sql("""
        UPDATE equipe_eventos
        SET titulo=?, data=?, local=?, descricao=?
        WHERE id=?
    """), (titulo, data_ev, local_ev, descricao, id))

    c.execute(fix_sql("DELETE FROM equipe_eventos_membros WHERE evento_id=?"), (id,))
    for membro in equipe:
        membro = str(membro).strip()
        if membro:
            c.execute(fix_sql("""
                INSERT INTO equipe_eventos_membros (evento_id, funcionario)
                VALUES (?, ?)
            """), (id, membro))

    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/equipes-eventos/<int:id>", methods=["DELETE"])
def equipes_eventos_deletar(id):
    if not session.get("nome"):
        return jsonify({"erro": "não logado"}), 401

    conn = get_conn()
    c = cursor(conn)
    c.execute(fix_sql("DELETE FROM equipe_eventos_membros WHERE evento_id=?"), (id,))
    c.execute(fix_sql("DELETE FROM equipe_eventos WHERE id=?"), (id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/equipes-eventos/funcionarios")
def equipes_eventos_funcionarios():
    if not session.get("nome"):
        return jsonify([]), 401
    try:
        df = pd.read_csv(FUNCIONARIOS_URL)
        df.columns = df.columns.str.strip()
        df = df.fillna("")
        nomes = df["NOME"].astype(str).str.strip().tolist()
        nomes = [n for n in nomes if n]
        return jsonify(nomes)
    except Exception as e:
        print("ERRO funcionarios:", e)
        return jsonify([])

@app.route("/atas")
def atas():
    if not session.get("nome"):
        return redirect("/")
    return render_template("atas.html", nome=session["nome"], cargo=session.get("cargo", "funcionario"))


@app.route("/atas/dados")
def atas_dados():
    if not session.get("nome"):
        return jsonify([]), 401
    conn = get_conn()
    c = cursor(conn)
    c.execute("""
        SELECT id, titulo, data_reuniao, presentes, pauta, deliberacoes, link, autor, criado_em
        FROM atas
        ORDER BY data_reuniao DESC, criado_em DESC
    """)
    rows = c.fetchall()
    conn.close()
    if is_pg():
        return jsonify([dict(r) for r in rows])
    return jsonify([{
        "id": r[0], "titulo": r[1], "data_reuniao": r[2], "presentes": r[3],
        "pauta": r[4], "deliberacoes": r[5], "link": r[6], "autor": r[7], "criado_em": r[8],
    } for r in rows])


@app.route("/atas", methods=["POST"])
def atas_criar():
    if not session.get("nome"):
        return jsonify({"erro": "não logado"}), 401
    data = request.json
    titulo = str(data.get("titulo", "")).strip()
    if not titulo:
        return jsonify({"erro": "Título obrigatório"}), 400
    conn = get_conn()
    c = cursor(conn)
    c.execute(fix_sql("""
        INSERT INTO atas (titulo, data_reuniao, presentes, pauta, deliberacoes, link, autor)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """), (
        titulo,
        data.get("data_reuniao", ""),
        data.get("presentes", ""),
        data.get("pauta", ""),
        data.get("deliberacoes", ""),
        data.get("link", ""),
        session["nome"]
    ))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/atas/<int:id>", methods=["DELETE"])
def atas_deletar(id):
    if not session.get("nome"):
        return jsonify({"erro": "não logado"}), 401
    conn = get_conn()
    c = cursor(conn)
    c.execute(fix_sql("SELECT autor FROM atas WHERE id=?"), (id,))
    row = c.fetchone()
    if not row:
        conn.close()
        return jsonify({"erro": "Ata não encontrada"}), 404
    autor = row["autor"] if is_pg() else row[0]
    if autor != session["nome"] and session.get("cargo") != "chefe":
        conn.close()
        return jsonify({"erro": "Sem permissão"}), 403
    c.execute(fix_sql("DELETE FROM atas WHERE id=?"), (id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/ramais")
def ramais():
    if not session.get("nome"):
        return redirect("/")
    return render_template(
        "ramais.html",
        nome=session["nome"],
        cargo=session.get("cargo", "funcionario")
    )


@app.route("/ramais/dados")
def ramais_dados():
    if not session.get("nome"):
        return jsonify([]), 401
    conn = get_conn()
    c = cursor(conn)
    c.execute("""
        SELECT id, nome, ramal, cargo, setor, criado_em
        FROM ramais
        ORDER BY setor, nome
    """)
    rows = c.fetchall()
    conn.close()
    if is_pg():
        return jsonify([dict(r) for r in rows])
    return jsonify([{
        "id": r[0], "nome": r[1], "ramal": r[2],
        "cargo": r[3], "setor": r[4], "criado_em": r[5],
    } for r in rows])


@app.route("/ramais", methods=["POST"])
def ramais_criar():
    if not session.get("nome"):
        return jsonify({"erro": "não logado"}), 401
    if session.get("cargo") != "chefe":
        return jsonify({"erro": "sem permissão"}), 403
    data = request.json
    nome  = str(data.get("nome", "")).strip()
    ramal = str(data.get("ramal", "")).strip()
    if not nome or not ramal:
        return jsonify({"erro": "Nome e ramal obrigatórios"}), 400
    conn = get_conn()
    c = cursor(conn)
    c.execute(fix_sql("""
        INSERT INTO ramais (nome, ramal, cargo, setor)
        VALUES (?, ?, ?, ?)
    """), (nome, ramal, data.get("cargo", ""), data.get("setor", "Geral")))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/ramais/<int:id>", methods=["DELETE"])
def ramais_deletar(id):
    if not session.get("nome"):
        return jsonify({"erro": "não logado"}), 401
    if session.get("cargo") != "chefe":
        return jsonify({"erro": "sem permissão"}), 403
    conn = get_conn()
    c = cursor(conn)
    c.execute(fix_sql("DELETE FROM ramais WHERE id=?"), (id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

def enviar_email(destinatario, assunto, corpo_html):
    def _enviar():
        try:
            response = requests.post(
                "https://api.resend.com/emails",
                headers={
                    "Authorization": f"Bearer {os.environ.get('RESEND_API_KEY')}",
                    "Content-Type": "application/json"
                },
                json={
                    "from": "Sistema CITE <onboarding@resend.dev>",
                    "to": [destinatario],
                    "subject": assunto,
                    "html": corpo_html
                }
            )
            print(f"✉️ Email enviado para {destinatario}: {response.status_code}")
        except Exception as e:
            print(f"⚠️ Falha ao enviar email: {e}")

    thread = threading.Thread(target=_enviar)
    thread.daemon = True
    thread.start()


def template_email(titulo, subtitulo, corpo, cor_destaque="#6366f1"):
    return f"""
    <!DOCTYPE html>
    <html lang="pt-br">
    <head>
      <meta charset="UTF-8">
      <meta name="viewport" content="width=device-width, initial-scale=1.0">
    </head>
    <body style="margin:0;padding:0;background:#f1f5f9;font-family:'Segoe UI',Arial,sans-serif;">
      <table width="100%" cellpadding="0" cellspacing="0" style="background:#f1f5f9;padding:40px 20px;">
        <tr><td align="center">
          <table width="100%" style="max-width:560px;background:white;border-radius:16px;overflow:hidden;box-shadow:0 4px 24px rgba(0,0,0,0.08);">
            <tr>
              <td style="background:linear-gradient(135deg,{cor_destaque},#818cf8);padding:32px 36px;">
                <div style="font-size:1.1rem;font-weight:800;color:white;margin-bottom:4px;">🌳 Sistema CITE</div>
                <div style="font-size:1.5rem;font-weight:800;color:white;margin-bottom:4px;">{titulo}</div>
                <div style="font-size:0.9rem;color:rgba(255,255,255,0.85);">{subtitulo}</div>
              </td>
            </tr>
            <tr>
              <td style="padding:32px 36px;">
                {corpo}
              </td>
            </tr>
            <tr>
              <td style="padding:20px 36px;border-top:1px solid #e2e8f0;background:#f8fafc;">
                <p style="margin:0;font-size:0.75rem;color:#94a3b8;text-align:center;">
                  Este é um email automático do Sistema CITE — não responda este email.<br>
                  Prefeitura Municipal de São José dos Campos
                </p>
              </td>
            </tr>
          </table>
        </td></tr>
      </table>
    </body>
    </html>
    """


def email_ci_confirmacao(email_func, nome_func, tipo, datas, obs):
    obs_linha = f"""
        <tr>
          <td style="padding:8px 0;font-size:0.85rem;color:#64748b;font-weight:600;">Observações</td>
          <td style="padding:8px 0;font-size:0.85rem;color:#1e293b;">{obs}</td>
        </tr>
    """ if obs else ""

    corpo = f"""
        <p style="margin:0 0 20px;font-size:0.95rem;color:#1e293b;">
            Olá, <strong>{nome_func}</strong>! Sua comunicação interna foi registrada com sucesso. ✅
        </p>
        <div style="background:#f8fafc;border-radius:12px;padding:20px 24px;margin-bottom:20px;border:1px solid #e2e8f0;">
          <table width="100%" cellpadding="0" cellspacing="0">
            <tr>
              <td style="padding:8px 0;font-size:0.85rem;color:#64748b;font-weight:600;width:140px;">Tipo</td>
              <td style="padding:8px 0;font-size:0.85rem;color:#1e293b;font-weight:700;">{tipo}</td>
            </tr>
            <tr>
              <td style="padding:8px 0;font-size:0.85rem;color:#64748b;font-weight:600;">Data(s)</td>
              <td style="padding:8px 0;font-size:0.85rem;color:#1e293b;">{datas}</td>
            </tr>
            {obs_linha}
          </table>
        </div>
        <p style="margin:0;font-size:0.85rem;color:#64748b;">
            Você será notificado por email quando a chefia registrar uma devolutiva.
        </p>
    """
    html = template_email(titulo="CI Registrada", subtitulo="Sua comunicação interna foi enviada", corpo=corpo)
    enviar_email(email_func, f"✅ CI registrada — {tipo}", html)


def email_ci_devolutiva(email_func, nome_func, tipo, datas, devolutiva):
    cores = {
        "Deferido":   ("#22c55e", "#dcfce7", "✅"),
        "Indeferido": ("#ef4444", "#fee2e2", "❌"),
        "Ciente":     ("#3b82f6", "#dbeafe", "🔵"),
    }
    cor, bg, emoji = cores.get(devolutiva, ("#6366f1", "#eef2ff", "📋"))

    corpo = f"""
        <p style="margin:0 0 20px;font-size:0.95rem;color:#1e293b;">
            Olá, <strong>{nome_func}</strong>! A chefia registrou uma devolutiva para sua CI.
        </p>
        <div style="background:{bg};border-radius:12px;padding:20px 24px;margin-bottom:20px;border-left:4px solid {cor};">
          <div style="font-size:1.1rem;font-weight:800;color:{cor};margin-bottom:8px;">{emoji} {devolutiva}</div>
          <table width="100%" cellpadding="0" cellspacing="0">
            <tr>
              <td style="padding:6px 0;font-size:0.85rem;color:#64748b;font-weight:600;width:140px;">Tipo da CI</td>
              <td style="padding:6px 0;font-size:0.85rem;color:#1e293b;font-weight:700;">{tipo}</td>
            </tr>
            <tr>
              <td style="padding:6px 0;font-size:0.85rem;color:#64748b;font-weight:600;">Data(s)</td>
              <td style="padding:6px 0;font-size:0.85rem;color:#1e293b;">{datas}</td>
            </tr>
          </table>
        </div>
        <p style="margin:0;font-size:0.85rem;color:#64748b;">Acesse o sistema para mais detalhes.</p>
    """
    html = template_email(titulo=f"CI {devolutiva}", subtitulo="A chefia registrou uma devolutiva para sua CI", corpo=corpo, cor_destaque=cor)
    enviar_email(email_func, f"{emoji} CI {devolutiva} — {tipo}", html)


def email_solicitacao_confirmacao(email_func, nome_func, tipo, descricao):
    desc_linha = f"""
        <tr>
          <td style="padding:8px 0;font-size:0.85rem;color:#64748b;font-weight:600;">Descrição</td>
          <td style="padding:8px 0;font-size:0.85rem;color:#1e293b;">{descricao}</td>
        </tr>
    """ if descricao else ""

    corpo = f"""
        <p style="margin:0 0 20px;font-size:0.95rem;color:#1e293b;">
            Olá, <strong>{nome_func}</strong>! Sua solicitação foi registrada e está aguardando análise da chefia. ⏳
        </p>
        <div style="background:#f8fafc;border-radius:12px;padding:20px 24px;margin-bottom:20px;border:1px solid #e2e8f0;">
          <table width="100%" cellpadding="0" cellspacing="0">
            <tr>
              <td style="padding:8px 0;font-size:0.85rem;color:#64748b;font-weight:600;width:140px;">Tipo</td>
              <td style="padding:8px 0;font-size:0.85rem;color:#1e293b;font-weight:700;">{tipo}</td>
            </tr>
            {desc_linha}
          </table>
        </div>
        <p style="margin:0;font-size:0.85rem;color:#64748b;">
            Você será notificado por email quando a chefia responder sua solicitação.
        </p>
    """
    html = template_email(titulo="Solicitação Registrada", subtitulo="Sua solicitação foi enviada para análise", corpo=corpo)
    enviar_email(email_func, f"📋 Solicitação registrada — {tipo}", html)


def email_solicitacao_resposta(email_func, nome_func, tipo, status, resposta):
    aprovada = status == "aprovada"
    cor  = "#22c55e" if aprovada else "#ef4444"
    bg   = "#dcfce7" if aprovada else "#fee2e2"
    emoji = "✅" if aprovada else "❌"
    label = "Aprovada" if aprovada else "Recusada"

    resposta_linha = f"""
        <tr>
          <td style="padding:8px 0;font-size:0.85rem;color:#64748b;font-weight:600;">Resposta</td>
          <td style="padding:8px 0;font-size:0.85rem;color:#1e293b;">{resposta}</td>
        </tr>
    """ if resposta else ""

    corpo = f"""
        <p style="margin:0 0 20px;font-size:0.95rem;color:#1e293b;">
            Olá, <strong>{nome_func}</strong>! A chefia respondeu sua solicitação.
        </p>
        <div style="background:{bg};border-radius:12px;padding:20px 24px;margin-bottom:20px;border-left:4px solid {cor};">
          <div style="font-size:1.1rem;font-weight:800;color:{cor};margin-bottom:8px;">{emoji} {label}</div>
          <table width="100%" cellpadding="0" cellspacing="0">
            <tr>
              <td style="padding:6px 0;font-size:0.85rem;color:#64748b;font-weight:600;width:140px;">Tipo</td>
              <td style="padding:6px 0;font-size:0.85rem;color:#1e293b;font-weight:700;">{tipo}</td>
            </tr>
            {resposta_linha}
          </table>
        </div>
        <p style="margin:0;font-size:0.85rem;color:#64748b;">Acesse o sistema para mais detalhes.</p>
    """
    html = template_email(titulo=f"Solicitação {label}", subtitulo="A chefia respondeu sua solicitação", corpo=corpo, cor_destaque=cor)
    enviar_email(email_func, f"{emoji} Solicitação {label.lower()} — {tipo}", html)

@app.route("/ping", methods=["POST"])
def ping():
    if session.get("nome"):
        session.modified = True
        return jsonify({"ok": True})
    return jsonify({"expirado": True}), 401


# =========================
# 📊 EXPORTAR EXCEL
# =========================
@app.route("/relatorios/excel/ci")
def exportar_excel_ci():
    if not session.get("nome"):
        return redirect("/")

    nome_func = request.args.get("funcionario", "").strip()
    if session.get("cargo") != "chefe":
        nome_func = session["nome"]

    resultado = []
    try:
        if nome_func:
            df_func = pd.read_csv(FUNCIONARIOS_URL)
            df_func.columns = df_func.columns.str.strip()
            df_func = df_func.fillna("")
            df_func["NOME"] = df_func["NOME"].astype(str).str.strip()
            row_func = df_func[df_func["NOME"] == nome_func]
            if not row_func.empty:
                url_csv = str(row_func.iloc[0].get("LINK_CSV", "")).strip()
                sheet_match = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url_csv)
                if sheet_match:
                    client = get_gspread_client()
                    sh = client.open_by_key(sheet_match.group(1))
                    ws = sh.worksheet("CI")
                    rows = ws.get_all_values()
                    for r in rows[1:]:
                        if len(r) < 4 or not r[0]:
                            continue
                        resultado.append({
                            "Carimbo":     r[0] if len(r) > 0 else "",
                            "Funcionário": r[2] if len(r) > 2 else "",
                            "Tipo":        r[3] if len(r) > 3 else "",
                            "Data(s)":     r[4] if len(r) > 4 else "",
                            "Observações": r[5] if len(r) > 5 else "",
                            "Devolutiva":  r[6] if len(r) > 6 else "",
                        })
    except Exception as e:
        print("ERRO exportar_excel_ci:", e)

    wb = openpyxl.Workbook()
    ws_excel = wb.active
    ws_excel.title = "CIs"

    header_fill = PatternFill("solid", fgColor="6366F1")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    alt_fill    = PatternFill("solid", fgColor="EEF2FF")
    border      = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    ws_excel.merge_cells("A1:F1")
    titulo_cell = ws_excel["A1"]
    titulo_cell.value = f"Relatório de CIs — {nome_func} — {datetime.now().strftime('%d/%m/%Y')}"
    titulo_cell.font  = Font(bold=True, size=13, color="1E293B")
    titulo_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_excel.row_dimensions[1].height = 28

    colunas = ["Carimbo", "Funcionário", "Tipo", "Data(s)", "Observações", "Devolutiva"]
    for col_idx, col_name in enumerate(colunas, 1):
        cell = ws_excel.cell(row=2, column=col_idx, value=col_name)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
    ws_excel.row_dimensions[2].height = 22

    for row_idx, row_data in enumerate(resultado, 3):
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, key in enumerate(colunas, 1):
            cell = ws_excel.cell(row=row_idx, column=col_idx, value=row_data.get(key, ""))
            cell.border    = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if fill:
                cell.fill = fill

    larguras = [22, 24, 22, 16, 30, 14]
    for i, w in enumerate(larguras, 1):
        ws_excel.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    from flask import send_file
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"CIs_{nome_func.replace(' ','_')}_{datetime.now().strftime('%Y%m')}.xlsx"
    )


@app.route("/relatorios/excel/solicitacoes")
def exportar_excel_solicitacoes():
    if not session.get("nome"):
        return redirect("/")

    conn = get_conn()
    c = cursor(conn)
    if session.get("cargo") == "chefe":
        c.execute("""
            SELECT solicitante, tipo, descricao, data_ref, data_pedido, status, resposta, respondido_em
            FROM solicitacoes ORDER BY data_pedido DESC
        """)
    else:
        c.execute(
            fix_sql("""
                SELECT solicitante, tipo, descricao, data_ref, data_pedido, status, resposta, respondido_em
                FROM solicitacoes WHERE solicitante=? ORDER BY data_pedido DESC
            """),
            (session["nome"],)
        )
    rows = c.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws_excel = wb.active
    ws_excel.title = "Solicitações"

    header_fill = PatternFill("solid", fgColor="6366F1")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    alt_fill    = PatternFill("solid", fgColor="EEF2FF")
    border      = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    ws_excel.merge_cells("A1:H1")
    titulo_cell = ws_excel["A1"]
    titulo_cell.value = f"Relatório de Solicitações — {datetime.now().strftime('%d/%m/%Y')}"
    titulo_cell.font  = Font(bold=True, size=13, color="1E293B")
    titulo_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_excel.row_dimensions[1].height = 28

    colunas = ["Solicitante", "Tipo", "Descrição", "Data Ref.", "Data Pedido", "Status", "Resposta", "Respondido em"]
    for col_idx, col_name in enumerate(colunas, 1):
        cell = ws_excel.cell(row=2, column=col_idx, value=col_name)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
    ws_excel.row_dimensions[2].height = 22

    for row_idx, row in enumerate(rows, 3):
        fill = alt_fill if row_idx % 2 == 0 else None
        vals = list(row.values()) if is_pg() else list(row)
        for col_idx, val in enumerate(vals, 1):
            cell = ws_excel.cell(row=row_idx, column=col_idx, value=val or "")
            cell.border    = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if fill:
                cell.fill = fill

    larguras = [24, 20, 30, 12, 18, 12, 30, 18]
    for i, w in enumerate(larguras, 1):
        ws_excel.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    from flask import send_file
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"Solicitacoes_{datetime.now().strftime('%Y%m')}.xlsx"
    )


# =========================
# 📄 RELATÓRIO MENSAL PDF
# =========================
def gerar_pdf_mensal(nome_chefe=None):
    buf   = BytesIO()
    doc   = SimpleDocTemplate(buf, pagesize=A4,
                               leftMargin=2*cm, rightMargin=2*cm,
                               topMargin=2*cm, bottomMargin=2*cm)
    story = []

    styles = getSampleStyleSheet()
    cor_primaria = colors.HexColor("#6366f1")

    st_titulo = ParagraphStyle("titulo", fontName="Helvetica-Bold", fontSize=18,
        textColor=cor_primaria, spaceAfter=4, alignment=TA_LEFT)
    st_sub = ParagraphStyle("sub", fontName="Helvetica", fontSize=10,
        textColor=colors.HexColor("#64748b"), spaceAfter=20)
    st_secao = ParagraphStyle("secao", fontName="Helvetica-Bold", fontSize=13,
        textColor=colors.HexColor("#1e293b"), spaceBefore=18, spaceAfter=8)
    st_normal = ParagraphStyle("normal", fontName="Helvetica", fontSize=9,
        textColor=colors.HexColor("#1e293b"), spaceAfter=4)

    mes_atual  = datetime.now().month
    ano_atual  = datetime.now().year
    MESES_PT   = ["", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
                  "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]

    story.append(Paragraph("🌳 Sistema CITE", st_titulo))
    story.append(Paragraph(
        f"Relatório Mensal — {MESES_PT[mes_atual]}/{ano_atual} &nbsp;|&nbsp; Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}",
        st_sub))
    story.append(HRFlowable(width="100%", thickness=2, color=cor_primaria, spaceAfter=16))

    story.append(Paragraph("📋 Solicitações do Mês", st_secao))

    conn = get_conn()
    c    = cursor(conn)

    if is_pg():
        c.execute("""
            SELECT solicitante, tipo, data_ref, data_pedido, status, resposta
            FROM solicitacoes
            WHERE EXTRACT(MONTH FROM data_pedido) = %s AND EXTRACT(YEAR FROM data_pedido) = %s
            ORDER BY data_pedido DESC
        """, (mes_atual, ano_atual))
    else:
        c.execute("""
            SELECT solicitante, tipo, data_ref, data_pedido, status, resposta
            FROM solicitacoes
            WHERE strftime('%m', data_pedido) = ? AND strftime('%Y', data_pedido) = ?
            ORDER BY data_pedido DESC
        """, (f"{mes_atual:02d}", str(ano_atual)))

    solic = c.fetchall()

    total_s   = len(solic)
    aprovadas = sum(1 for s in solic if (s["status"] if is_pg() else s[4]) == "aprovada")
    recusadas = sum(1 for s in solic if (s["status"] if is_pg() else s[4]) == "recusada")
    pendentes = sum(1 for s in solic if (s["status"] if is_pg() else s[4]) == "pendente")

    resumo_data = [
        ["Total", "Aprovadas", "Recusadas", "Pendentes"],
        [str(total_s), str(aprovadas), str(recusadas), str(pendentes)],
    ]
    t_resumo = Table(resumo_data, colWidths=[4*cm]*4)
    t_resumo.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), cor_primaria),
        ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
        ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0), (-1,-1), 9),
        ("ALIGN",      (0,0), (-1,-1), "CENTER"),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ("ROWHEIGHT",  (0,0), (-1,-1), 18),
        ("BACKGROUND", (0,1), (-1,1), colors.HexColor("#eef2ff")),
        ("GRID",       (0,0), (-1,-1), 0.5, colors.HexColor("#e2e8f0")),
    ]))
    story.append(t_resumo)
    story.append(Spacer(1, 10))

    if solic:
        dados_s = [["Solicitante", "Tipo", "Data", "Status", "Resposta"]]
        for s in solic:
            if is_pg():
                dados_s.append([s["solicitante"] or "", s["tipo"] or "",
                    str(s["data_pedido"])[:10] if s["data_pedido"] else "",
                    s["status"] or "", (s["resposta"] or "")[:40]])
            else:
                dados_s.append([s[0] or "", s[1] or "", s[3][:10] if s[3] else "",
                    s[4] or "", (s[5] or "")[:40]])
        t_solic = Table(dados_s, colWidths=[4*cm, 3.5*cm, 2.5*cm, 2.5*cm, 4.5*cm])
        t_solic.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1e293b")),
            ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
            ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTNAME",   (0,1), (-1,-1), "Helvetica"),
            ("FONTSIZE",   (0,0), (-1,-1), 8),
            ("ALIGN",      (0,0), (-1,-1), "LEFT"),
            ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
            ("ROWHEIGHT",  (0,0), (-1,-1), 16),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#f8fafc")]),
            ("GRID",       (0,0), (-1,-1), 0.5, colors.HexColor("#e2e8f0")),
            ("LEFTPADDING",  (0,0), (-1,-1), 6),
            ("RIGHTPADDING", (0,0), (-1,-1), 6),
        ]))
        story.append(t_solic)
    else:
        story.append(Paragraph("Nenhuma solicitação no mês.", st_normal))

    story.append(Spacer(1, 16))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#e2e8f0"), spaceAfter=8))
    story.append(Paragraph("📊 Resumo Geral", st_secao))

    if is_pg():
        c.execute("SELECT COUNT(*) FROM solicitacoes WHERE EXTRACT(MONTH FROM data_pedido)=%s AND EXTRACT(YEAR FROM data_pedido)=%s",
                  (mes_atual, ano_atual))
    else:
        c.execute("SELECT COUNT(*) FROM solicitacoes WHERE strftime('%m',data_pedido)=? AND strftime('%Y',data_pedido)=?",
                  (f"{mes_atual:02d}", str(ano_atual)))
    total_sol_row = c.fetchone()
    total_sol = list(total_sol_row.values())[0] if is_pg() else total_sol_row[0]

    if is_pg():
        c.execute("SELECT COUNT(DISTINCT solicitante) FROM solicitacoes WHERE EXTRACT(MONTH FROM data_pedido)=%s AND EXTRACT(YEAR FROM data_pedido)=%s",
                  (mes_atual, ano_atual))
    else:
        c.execute("SELECT COUNT(DISTINCT solicitante) FROM solicitacoes WHERE strftime('%m',data_pedido)=? AND strftime('%Y',data_pedido)=?",
                  (f"{mes_atual:02d}", str(ano_atual)))
    func_row = c.fetchone()
    func_ativos = list(func_row.values())[0] if is_pg() else func_row[0]

    conn.close()

    resumo_geral = [
        ["Métrica", "Valor"],
        ["Total de Solicitações no Mês", str(total_sol)],
        ["Funcionários com Solicitações", str(func_ativos)],
        ["Mês de Referência", f"{MESES_PT[mes_atual]}/{ano_atual}"],
        ["Relatório Gerado em", datetime.now().strftime("%d/%m/%Y %H:%M")],
        ["Gerado por", nome_chefe or "Sistema"],
    ]
    t_geral = Table(resumo_geral, colWidths=[10*cm, 7*cm])
    t_geral.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), cor_primaria),
        ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
        ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTNAME",   (0,1), (-1,-1), "Helvetica"),
        ("FONTSIZE",   (0,0), (-1,-1), 9),
        ("ALIGN",      (0,0), (0,-1), "LEFT"),
        ("ALIGN",      (1,0), (1,-1), "CENTER"),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ("ROWHEIGHT",  (0,0), (-1,-1), 18),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#f8fafc")]),
        ("GRID",       (0,0), (-1,-1), 0.5, colors.HexColor("#e2e8f0")),
        ("LEFTPADDING",  (0,0), (-1,-1), 8),
    ]))
    story.append(t_geral)

    doc.build(story)
    buf.seek(0)
    return buf


@app.route("/relatorios/pdf")
def exportar_pdf():
    if not session.get("nome") or session.get("cargo") != "chefe":
        return redirect("/")
    from flask import send_file
    buf = gerar_pdf_mensal(nome_chefe=session["nome"])
    mes = datetime.now().strftime("%Y%m")
    return send_file(buf, mimetype="application/pdf", as_attachment=True,
                     download_name=f"Relatorio_Mensal_{mes}.pdf")


# =========================
# 📧 RELATÓRIO SEMANAL
# =========================
def gerar_html_relatorio_semanal():
    hoje      = datetime.now()
    semana_ini = (hoje - timedelta(days=7)).strftime("%Y-%m-%d")

    conn = get_conn()
    c    = cursor(conn)

    c.execute(fix_sql("""
        SELECT solicitante, tipo, status, data_pedido
        FROM solicitacoes
        WHERE data_pedido >= ?
        ORDER BY data_pedido DESC
    """), (semana_ini,))
    solic = c.fetchall()
    conn.close()

    total_s   = len(solic)
    aprovadas = sum(1 for s in solic if (s["status"] if is_pg() else s[2]) == "aprovada")
    recusadas = sum(1 for s in solic if (s["status"] if is_pg() else s[2]) == "recusada")
    pendentes = sum(1 for s in solic if (s["status"] if is_pg() else s[2]) == "pendente")

    linhas_solic = ""
    for s in solic[:10]:
        st = s["status"] if is_pg() else s[2]
        sl = s["solicitante"] if is_pg() else s[0]
        tp = s["tipo"] if is_pg() else s[1]
        dp = str(s["data_pedido"] if is_pg() else s[3])
        status_cor = {"aprovada": "#22c55e", "recusada": "#ef4444", "pendente": "#f59e0b"}.get(st, "#64748b")
        linhas_solic += f"""
        <tr>
          <td style="padding:8px 12px;font-size:0.82rem;border-bottom:1px solid #e2e8f0;">{sl}</td>
          <td style="padding:8px 12px;font-size:0.82rem;border-bottom:1px solid #e2e8f0;">{tp}</td>
          <td style="padding:8px 12px;font-size:0.82rem;border-bottom:1px solid #e2e8f0;">
            <span style="background:{status_cor}22;color:{status_cor};padding:2px 8px;border-radius:20px;font-weight:700;font-size:0.75rem;">{st.capitalize()}</span>
          </td>
          <td style="padding:8px 12px;font-size:0.82rem;border-bottom:1px solid #e2e8f0;color:#64748b;">{dp[:10]}</td>
        </tr>"""

    rodape_mais = f'<p style="text-align:center;font-size:0.8rem;color:#64748b;margin-top:8px;">+ {total_s - 10} mais não exibidos</p>' if total_s > 10 else ""
    periodo = f"{(hoje - timedelta(days=7)).strftime('%d/%m')} a {hoje.strftime('%d/%m/%Y')}"

    corpo = f"""
        <p style="margin:0 0 20px;font-size:0.95rem;color:#1e293b;">
            Olá! Aqui está o resumo da semana de <strong>{periodo}</strong>.
        </p>
        <table width="100%" cellpadding="0" cellspacing="0" style="margin-bottom:24px;">
          <tr>
            <td width="25%" style="padding:0 6px 0 0;">
              <div style="background:#eef2ff;border-radius:10px;padding:16px;text-align:center;">
                <div style="font-size:1.8rem;font-weight:800;color:#6366f1;">{total_s}</div>
                <div style="font-size:0.72rem;color:#64748b;font-weight:600;margin-top:4px;">SOLICITAÇÕES</div>
              </div>
            </td>
            <td width="25%" style="padding:0 6px;">
              <div style="background:#dcfce7;border-radius:10px;padding:16px;text-align:center;">
                <div style="font-size:1.8rem;font-weight:800;color:#22c55e;">{aprovadas}</div>
                <div style="font-size:0.72rem;color:#64748b;font-weight:600;margin-top:4px;">APROVADAS</div>
              </div>
            </td>
            <td width="25%" style="padding:0 6px;">
              <div style="background:#fee2e2;border-radius:10px;padding:16px;text-align:center;">
                <div style="font-size:1.8rem;font-weight:800;color:#ef4444;">{recusadas}</div>
                <div style="font-size:0.72rem;color:#64748b;font-weight:600;margin-top:4px;">RECUSADAS</div>
              </div>
            </td>
            <td width="25%" style="padding:0 0 0 6px;">
              <div style="background:#fef3c7;border-radius:10px;padding:16px;text-align:center;">
                <div style="font-size:1.8rem;font-weight:800;color:#f59e0b;">{pendentes}</div>
                <div style="font-size:0.72rem;color:#64748b;font-weight:600;margin-top:4px;">PENDENTES</div>
              </div>
            </td>
          </tr>
        </table>
        {"" if not solic else f'''
        <div style="margin-bottom:8px;font-size:0.85rem;font-weight:700;color:#1e293b;">📋 Solicitações da Semana</div>
        <table width="100%" cellpadding="0" cellspacing="0" style="border:1px solid #e2e8f0;border-radius:10px;overflow:hidden;">
          <tr style="background:#f8fafc;">
            <th style="padding:10px 12px;font-size:0.78rem;color:#64748b;text-align:left;font-weight:700;">Funcionário</th>
            <th style="padding:10px 12px;font-size:0.78rem;color:#64748b;text-align:left;font-weight:700;">Tipo</th>
            <th style="padding:10px 12px;font-size:0.78rem;color:#64748b;text-align:left;font-weight:700;">Status</th>
            <th style="padding:10px 12px;font-size:0.78rem;color:#64748b;text-align:left;font-weight:700;">Data</th>
          </tr>
          {linhas_solic}
        </table>
        {rodape_mais}
        '''}
        <p style="margin:20px 0 0;font-size:0.85rem;color:#64748b;">
            Acesse o sistema para mais detalhes ou para gerar o relatório completo em PDF.
        </p>
    """

    return template_email(titulo="Relatório Semanal", subtitulo=f"Resumo de {periodo}", corpo=corpo)


def enviar_relatorio_semanal():
    print(f"📧 Enviando relatório semanal — {datetime.now()}")
    try:
        df = pd.read_csv(FUNCIONARIOS_URL)
        df.columns = df.columns.str.strip()
        df = df.fillna("")
        chefes = df[df["CARGO"].str.strip().str.lower() == "chefe"]

        html = gerar_html_relatorio_semanal()

        for _, row in chefes.iterrows():
            email_chefe = str(row.get("EMAIL INSTITUCIONAL", "")).strip()
            nome_chefe  = str(row.get("NOME", "")).strip()
            if email_chefe:
                enviar_email(email_chefe, "📊 Relatório Semanal — Sistema CITE", html)
                print(f"  ✉️ Enviado para {nome_chefe} ({email_chefe})")
    except Exception as e:
        print(f"⚠️ Erro no relatório semanal: {e}")


@app.route("/relatorios/email-semanal/agora", methods=["POST"])
def relatorio_semanal_agora():
    if not session.get("nome") or session.get("cargo") != "chefe":
        return jsonify({"erro": "sem permissão"}), 403
    try:
        enviar_relatorio_semanal()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"erro": str(e)}), 500


# =========================
# ⏰ SCHEDULER
# =========================
scheduler = BackgroundScheduler(timezone="America/Sao_Paulo")
scheduler.add_job(
    enviar_relatorio_semanal,
    trigger="cron",
    day_of_week="mon",
    hour=8,
    minute=0
)
scheduler.start()
atexit.register(lambda: scheduler.shutdown())


# =========================
# 🌐 RELATÓRIOS
# =========================
@app.route("/relatorios")
def relatorios():
    if not session.get("nome") or session.get("cargo") != "chefe":
        return redirect("/dashboard")
    try:
        df = pd.read_csv(FUNCIONARIOS_URL)
        df.columns = df.columns.str.strip()
        df = df.fillna("")
        funcionarios = df["NOME"].astype(str).str.strip().tolist()
        funcionarios = [f for f in funcionarios if f]
    except:
        funcionarios = []
    return render_template(
        "relatorios.html",
        nome=session["nome"],
        cargo=session.get("cargo", "funcionario"),
        funcionarios=funcionarios
    )

@app.route('/favicon.ico')
def favicon():
    return send_from_directory('static', 'favicon.ico')

@app.route("/debug/sgrh")
def debug_sgrh():
    try:
        client = get_gspread_client()
        cred_json = os.environ.get("GOOGLE_CREDENTIALS")
        cred_dict = json.loads(cred_json)
        return jsonify({
            "client_email": cred_dict.get("client_email"),
            "ok": True
        })
    except Exception as e:
        return jsonify({"erro": str(e)})

# =========================
# 🚪 LOGOUT
# =========================
@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")


if __name__ == "__main__":
    init_db()
    app.run(debug=True)

# Garante init_db ao rodar com gunicorn
with app.app_context():
    init_db()